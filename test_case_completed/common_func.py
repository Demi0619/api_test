import openpyxl

def read_excel(excel_path,sheet_name):
    excel = openpyxl.load_workbook(excel_path)
    sheet = excel[sheet_name]
    tuple_list = []
    for value in sheet.values:
        if type(value[0]) is int:
            tuple_list.append(value)
    return tuple_list

def edit_excel_and_save(excel_path,sheet_name,row,column,value):
    excel = openpyxl.load_workbook(excel_path)
    sheet = excel[sheet_name]
    sheet.cell(row,column).value=value
    excel.save(excel_path)